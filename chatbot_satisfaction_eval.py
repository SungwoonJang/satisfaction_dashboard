"""
CLI 실행 모드: input 폴더에서 파일 입력 → 분석 → output 폴더에 엑셀 저장.
웹 대시보드는 app.py를 통해 실행하세요 (02_run_dashboard.bat).
"""

import os
import argparse

import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from core.categories import SESSION_COL, TIMESTAMP_COL, USER_ID_COL, SENDER_COL, TEXT_COL
from core.llm_client import DEFAULT_API_KEY, DEFAULT_MAX_WORKERS
from core.evaluator import run_pipeline, normalize_schema


# ──────────────────────────────────────────────────────────────────────────────
# 엑셀 내보내기 (기존 스타일 유지)
# ──────────────────────────────────────────────────────────────────────────────

def export_to_excel(result: dict, output_path: str):
    """run_pipeline() 결과 dict → 기존과 동일한 구조의 엑셀 저장."""
    detail_df         = result["detail_df"]
    # 엑셀 출력용 컬럼 (has_* 플래그 제외)
    out_cols = [c for c in detail_df.columns
                if c not in ("has_profanity","has_negative","has_gratitude","asks_counselor","agent_limit")]
    detail_out        = detail_df[out_cols].copy()
    satisfaction_df   = detail_out[detail_out["평가결과"] == "만족"].copy()
    dissatisfaction_df = detail_out[detail_out["평가결과"] == "불만족"].copy()

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    group_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    col_width_map = {
        TIMESTAMP_COL: 20, SESSION_COL: 15, USER_ID_COL: 15,
        TEXT_COL: 80, "대화요약": 40, "대분류": 15, "중분류": 15,
        "평가결과": 10, "평가이유": 60,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail_out.to_excel(writer, sheet_name="detail", index=False)
        satisfaction_df.to_excel(writer, sheet_name="satisfaction", index=False)
        dissatisfaction_df.to_excel(writer, sheet_name="dissatisfaction", index=False)
        result["main_category_stats"].to_excel(writer, sheet_name="summary_대분류통계", index=False)
        result["top10_dissatisfied"].to_excel(writer, sheet_name="summary_불만_top10", index=False)
        result["top10_satisfied"].to_excel(writer, sheet_name="summary_긍정_top10", index=False)

        # ── 데이터 시트 스타일 ──────────────────────────────────────────────
        for sheet_name, cur_df in [
            ("detail",          detail_out),
            ("satisfaction",    satisfaction_df),
            ("dissatisfaction", dissatisfaction_df),
        ]:
            if sheet_name not in writer.sheets:
                continue
            ws = writer.sheets[sheet_name]

            for col_idx, col_name in enumerate(cur_df.columns, start=1):
                if col_name in col_width_map:
                    ws.column_dimensions[get_column_letter(col_idx)].width = col_width_map[col_name]

            try:
                sess_idx = list(cur_df.columns).index(SESSION_COL)
            except ValueError:
                continue

            target_sessions = set(cur_df[SESSION_COL].unique()[::2])
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                apply_fill = row_idx > 1 and row[sess_idx].value in target_sessions
                for cell in row:
                    cell.border = thin
                    if apply_fill:
                        cell.fill = group_fill

        # ── 통계 시트 스타일 ────────────────────────────────────────────────
        for sheet_name in ["summary_대분류통계", "summary_불만_top10", "summary_긍정_top10"]:
            if sheet_name not in writer.sheets:
                continue
            ws = writer.sheets[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                if sheet_name in ("summary_불만_top10", "summary_긍정_top10"):
                    ws.column_dimensions[letter].width = {4: 50, 5: 70}.get(col_idx, 18)
                else:
                    ws.column_dimensions[letter].width = 18

        # ── Top10 시트 하단에 종합평가 추가 ────────────────────────────────
        def _append(ws, start_row, blocks):
            r = start_row
            for title, body in blocks:
                ws.cell(row=r, column=1, value=title).border = thin
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
                c = ws.cell(row=r, column=2, value=body)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                for col in range(1, 6):
                    ws.cell(row=r, column=col).border = thin
                ws.row_dimensions[r].height = min(15 + (len(body) // 40 + 1) * 15, 400)
                r += 2

        n_diss = len(result["top10_dissatisfied"])
        n_sat  = len(result["top10_satisfied"])
        if "summary_불만_top10" in writer.sheets:
            _append(writer.sheets["summary_불만_top10"], n_diss + 3 if n_diss else 3,
                    [("종합평가", result["diss_synthesis"])])
        if "summary_긍정_top10" in writer.sheets:
            sat = result["sat_llm_result"]
            _append(writer.sheets["summary_긍정_top10"], n_sat + 3 if n_sat else 3, [
                ("마케팅 메시지(50자 이내)", sat.get("marketing_message", "")),
                ("종합평가",                 sat.get("synthesis", "")),
            ])

    print(f"완료. 결과 파일: {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# CLI 진입점
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="챗봇 만족도 CLI 평가")
    parser.add_argument("input_file",             type=str,  help="input 폴더 내 파일명")
    parser.add_argument("--input_dir",  default="input",  type=str)
    parser.add_argument("--output_dir", default="output", type=str)
    parser.add_argument("--use_llm",    default="0",      type=str, help="1=LLM 사용, 0=Rule Only")
    args = parser.parse_args()

    use_llm    = args.use_llm in ("1", "true", "True", "YES", "yes", "Y", "y")
    input_path = os.path.join(args.input_dir, args.input_file)

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_path}")

    os.makedirs(args.output_dir, exist_ok=True)
    print(f"입력: {input_path}")
    print(f"출력: {args.output_dir}")
    print(f"LLM:  {'사용' if use_llm else '미사용 (Rule-Only)'}")

    ext = os.path.splitext(input_path)[1].lower()
    df  = pd.read_csv(input_path) if ext == ".csv" else pd.read_excel(input_path)
    df  = normalize_schema(df)
    df  = df.sort_values([SESSION_COL, TIMESTAMP_COL]).reset_index(drop=True)

    result = run_pipeline(df, use_llm=use_llm, api_key=DEFAULT_API_KEY, max_workers=DEFAULT_MAX_WORKERS)

    base_name   = os.path.splitext(os.path.basename(args.input_file))[0]
    output_path = os.path.join(args.output_dir, f"{base_name}_eval_result.xlsx")
    export_to_excel(result, output_path)
